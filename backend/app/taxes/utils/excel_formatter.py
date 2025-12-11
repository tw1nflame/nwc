import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment
from openpyxl.utils import get_column_letter
from typing import Dict, Union

def save_dataframes_to_excel(
    dataframes_dict: Dict[str, pd.DataFrame],
    output_path: str,
    date_format: str = "dd.mm.yy",
    header_height: int = 40
) -> None:
    """
    Сохраняет словарь датафреймов в форматированный Excel-файл.
    
    Параметры:
    dataframes_dict -- словарь {название_листа: датафрейм}
    output_path -- путь для сохранения Excel-файла
    date_format -- формат даты: "dd.mm.yy" (по умолчанию) или "month_year"
    header_height -- высота строки заголовка в пунктах (по умолчанию 40)
    """
    
    # Запись данных в Excel без форматирования
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Применение форматирования
    wb = load_workbook(output_path)

    header_alignment = Alignment(
        horizontal='center',
        vertical='center',
        wrap_text=True
    )
    
    for sheet_name, df in dataframes_dict.items():
        ws = wb[sheet_name]

        # Устанавливаем высоту строки для заголовков
        ws.row_dimensions[1].height = header_height

        # Применяем стиль переноса ко всем ячейкам заголовка
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.alignment = header_alignment
        
        # Определение типов столбцов
        date_cols = []
        pct_cols = []
        num_cols = []
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_dtype = df[col_name].dtype
            
            # Определение столбцов с датой
            if pd.api.types.is_datetime64_any_dtype(df[col_name]):
                date_cols.append(col_idx)
            
            # Определение процентных столбцов
            elif '%' in col_name:
                pct_cols.append(col_idx)
            
            # Определение числовых столбцов
            elif pd.api.types.is_numeric_dtype(col_dtype):
                num_cols.append(col_idx)
        
        # Применение форматирования к столбцам
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                col_idx = cell.column
                
                # Форматирование дат
                if col_idx in date_cols and cell.value:
                    if date_format == "month_year":
                        cell.number_format = 'MMMM YYYY'
                    else:
                        cell.number_format = numbers.FORMAT_DATE_DDMMYY
                
                # Форматирование процентов
                elif col_idx in pct_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = numbers.FORMAT_PERCENTAGE_00
                
                # Форматирование чисел
                elif col_idx in num_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = '### ### ### ### ##0'
        
        # Установка ширины столбцов
        for col_letter in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_letter)].width = 15
        
        # Добавление фильтров
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    
    # Сохранение результата
    wb.save(output_path)
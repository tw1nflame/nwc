import numpy as np
from io import BytesIO
import os
import pandas as pd
import psycopg2
from sqlalchemy import create_engine
import logging
from dotenv import load_dotenv
from utils.column_mapping import REVERSE_COLUMN_MAPPING, COLUMN_MAPPING
from utils.config import load_config
from utils.common import setup_custom_logging

load_dotenv()

# Создаем директорию для логов
log_dir = os.path.join(os.path.dirname(__file__), '..', 'logs')
os.makedirs(log_dir, exist_ok=True)

# Инициализируем логгер
setup_custom_logging(os.path.join(log_dir, "db_operations.log"))
logger = logging.getLogger(__name__)

DB_HOST = os.getenv('DB_HOST')
DB_PORT = int(os.getenv('DB_PORT', 5432))
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')
DB_NAME = os.getenv('DB_NAME')

DATA_TABLE = os.getenv('DATA_TABLE')
ADJUSTMENTS_TABLE = os.getenv('ADJUSTMENTS_TABLE')
EXCHANGE_RATE_TABLE = os.getenv('EXCHANGE_RATE_TABLE')
COEFFS_WITH_INTERCEPT_TABLE = os.getenv('COEFFS_WITH_INTERCEPT_TABLE')
COEFFS_NO_INTERCEPT_TABLE = os.getenv('COEFFS_NO_INTERCEPT_TABLE')
ENSEMBLE_INFO_TABLE = os.getenv('ENSEMBLE_INFO_TABLE')
BASEPLUS_TABULAR_ENSEMBLE_INFO_TABLE = os.getenv('BASEPLUS_TABULAR_ENSEMBLE_INFO_TABLE')
BASEPLUS_FEATURE_IMPORTANCE_TABLE = os.getenv('BASEPLUS_FEATURE_IMPORTANCE_TABLE')

SHEET_TO_TABLE = {
    'data': DATA_TABLE,
    'coeffs_with_intercept': COEFFS_WITH_INTERCEPT_TABLE,
    'coeffs_no_intercept': COEFFS_NO_INTERCEPT_TABLE,
    'TimeSeries_ensemble_models_info': ENSEMBLE_INFO_TABLE,
    'Tabular_ensemble_models_info': BASEPLUS_TABULAR_ENSEMBLE_INFO_TABLE,
    'Tabular_feature_importance': BASEPLUS_FEATURE_IMPORTANCE_TABLE,
}


CONFIG_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), '../../config_refined.yaml'))

def upload_pipeline_result_to_db(file_path: str, sheet_to_table: dict, date_value: str = None, date_column: str = None, pipeline: str = None):
    """
    Загружает все листы результата pipeline (BASE или BASE+) в соответствующие таблицы PostgreSQL через psycopg2.
    Если листы Tabular_ensemble_models_info или Tabular_feature_importance отсутствуют, 
    удаляет данные по указанной дате из соответствующих таблиц.
    
    Args:
        file_path: путь к Excel файлу
        sheet_to_table: маппинг листов к таблицам
        date_value: дата для удаления данных из отсутствующих листов (формат 'YYYY-MM-DD')
        date_column: имя столбца с датой в Excel (например, 'Дата'), будет преобразовано в имя столбца БД
    """
    logger.info(f"Начало загрузки результатов пайплайна в БД из файла: {file_path}")
    logger.info(f"Параметры: date_value={date_value}, date_column={date_column}")
    
    xls = pd.read_excel(file_path, sheet_name=None)
    logger.info(f"Загружены листы Excel: {list(xls.keys())}")
    
    # Создаем SQLAlchemy engine для pandas
    db_url = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    engine = create_engine(db_url)
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    logger.info(f"Подключение к БД установлено: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    
    try:
        cur = conn.cursor()
        
        # Получаем список статей, которые были обработаны (из листа 'data')
        processed_articles = []
        if 'data' in xls and not xls['data'].empty:
            # Применяем маппинг к колонкам для определения статей
            data_df = xls['data'].copy()
            data_df.columns = [COLUMN_MAPPING.get(col, col) for col in data_df.columns]
            if 'article' in data_df.columns:
                processed_articles = data_df['article'].unique().tolist()
                logger.info(f"Обработанные статьи: {processed_articles}")
            else:
                logger.warning("Колонка 'article' не найдена в данных после применения маппинга")
        else:
            logger.warning("Лист 'data' отсутствует или пуст")
        
        # Селективно удаляем данные только по обработанным статьям
        deleted_rows_total = 0
        for sheet, df in xls.items():
            table = sheet_to_table.get(sheet)
            if table and processed_articles:
                # Проверяем, есть ли в таблице колонка article
                cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s AND column_name = 'article'", (table,))
                has_article = cur.fetchone() is not None
                # Проверяем, есть ли колонка pipeline
                has_pipeline = False
                if pipeline:
                    cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s AND column_name = 'pipeline'", (table,))
                    has_pipeline = cur.fetchone() is not None
                if has_article:
                    placeholders = ', '.join(['%s'] * len(processed_articles))
                    if pipeline and has_pipeline:
                        delete_query = f'DELETE FROM {table} WHERE "article" IN ({placeholders}) AND pipeline = %s'
                        cur.execute(delete_query, processed_articles + [pipeline])
                    else:
                        delete_query = f'DELETE FROM {table} WHERE "article" IN ({placeholders})'
                        cur.execute(delete_query, processed_articles)
                    deleted_rows = cur.rowcount
                    deleted_rows_total += deleted_rows
                    logger.info(f"Удалено {deleted_rows} строк из таблицы {table} для статей: {processed_articles} (pipeline={pipeline if has_pipeline else 'any'})")
                else:
                    # Если нет колонки article, очищаем всю таблицу (для совместимости)
                    cur.execute(f'TRUNCATE TABLE {table} RESTART IDENTITY CASCADE')
                    logger.info(f"Таблица {table} полностью очищена (нет колонки article)")
            elif table:
                logger.info(f"Таблица {table} пропущена (нет обработанных статей)")
                
        logger.info(f"Всего удалено строк: {deleted_rows_total}")
        
        # Загружаем новые данные из Excel
        inserted_rows_total = 0
        for sheet, df in xls.items():
            table = sheet_to_table.get(sheet)
            if table is None:
                logger.warning(f"Лист '{sheet}' не найден в маппинге таблиц")
                continue
            if df.empty:
                logger.warning(f"Лист '{sheet}' пуст")
                continue

            logger.info(f"Обработка листа '{sheet}' -> таблица '{table}', строк: {len(df)}")
            logger.info(f"Форма DataFrame: {df.shape} (строки x колонки)")
            logger.info(f"Исходные колонки ({len(df.columns)}): {list(df.columns)}")

            # Применяем маппинг колонок из Excel в БД
            original_columns = df.columns.tolist()
            df.columns = [COLUMN_MAPPING.get(col, col) for col in df.columns]
            mapped_columns = df.columns.tolist()
            logger.debug(f"Маппинг колонок для {sheet}: {dict(zip(original_columns, mapped_columns))}")
            logger.info(f"Колонки после маппинга ({len(df.columns)}): {list(df.columns)}")

            # Проверяем, есть ли колонка pipeline в таблице
            has_pipeline = False
            if pipeline:
                cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s AND column_name = 'pipeline'", (table,))
                has_pipeline = cur.fetchone() is not None
            df_to_insert = df.copy()
            if pipeline and has_pipeline:
                if 'pipeline' not in df_to_insert.columns:
                    df_to_insert['pipeline'] = pipeline
                else:
                    df_to_insert['pipeline'] = pipeline
            columns = ', '.join([f'"{col}"' for col in df_to_insert.columns])
            values_template = ', '.join(['%s'] * len(df_to_insert.columns))
            insert_query = f'INSERT INTO {table} ({columns}) VALUES ({values_template})'

            logger.info(f"SQL запрос: {insert_query}")
            logger.info(f"Ожидается {len(df_to_insert.columns)} значений, колонки: {list(df_to_insert.columns)}")

            records = df_to_insert.values.tolist()
            logger.info(f"Первые 3 строки данных:")
            for i, row in enumerate(records[:3]):
                logger.info(f"  Строка {i+1}: {len(row)} значений - {row}")

            inserted_rows = 0
            for i, row in enumerate(records):
                try:
                    if len(row) != len(df_to_insert.columns):
                        logger.error(f"Несоответствие количества значений в строке {i+1}:")
                        logger.error(f"  Ожидается: {len(df_to_insert.columns)} значений")
                        logger.error(f"  Получено: {len(row)} значений")
                        logger.error(f"  Колонки ({len(df_to_insert.columns)}): {list(df_to_insert.columns)}")
                        logger.error(f"  Значения ({len(row)}): {row}")
                        raise ValueError(f"Количество значений ({len(row)}) не соответствует количеству колонок ({len(df_to_insert.columns)})")

                    cur.execute(insert_query, row)
                    inserted_rows += 1
                except Exception as e:
                    logger.error(f"Ошибка вставки строки {i+1} в таблицу {table}: {e}")
                    logger.error(f"Данные строки: {row}")
                    raise

            inserted_rows_total += inserted_rows
            logger.info(f"Вставлено {inserted_rows} строк в таблицу {table}")
        
        logger.info(f"Всего вставлено строк: {inserted_rows_total}")
        conn.commit()
        logger.info("Транзакция успешно завершена")
        
        
        cur.close()
        logger.info("Загрузка результатов пайплайна в БД завершена успешно")
        
    except Exception as e:
        logger.error(f"Ошибка при загрузке данных в БД: {e}", exc_info=True)
        conn.rollback()
        logger.info("Транзакция отменена")
        raise e
    finally:
        conn.close()
        logger.info("Соединение с БД закрыто")

def set_pipeline_column(date_column: str, date_value: str, pipeline_value: str, articles_processed: list = None):
    """
    Устанавливает значение pipeline в DATA_TABLE для записей с указанной датой.
    Если указан список статей, обновляет только эти статьи.
    date_column: имя столбца с датой в Excel (например, 'Дата'), будет преобразовано в имя столбца БД
    date_value: значение даты прогноза (например, '2025-01-01')
    pipeline_value: 'BASE' или 'BASE+'
    articles_processed: список статей, которые были обработаны (опционально)
    """
    logger = setup_custom_logging()
    logger.info(f"Установка pipeline колонки: прогноз на {date_value} -> {pipeline_value}")
    if articles_processed:
        logger.info(f"Обрабатываемые статьи: {articles_processed}")
    
    table = DATA_TABLE
    # Преобразуем имя столбца из Excel в имя столбца БД
    db_date_column = COLUMN_MAPPING.get(date_column, date_column)
    logger.info(f"Маппинг колонки: {date_column} -> {db_date_column}")

    # Load config to get USD articles
    config = load_config(CONFIG_PATH)
    usd_articles = config.get('Статьи для предикта в USD', [])

    # Prepare articles_processed with _USD if needed
    articles_for_update = None
    if articles_processed:
        articles_for_update = [
            (a + '_USD') if a in usd_articles else a
            for a in articles_processed
        ]
        logger.info(f"Статьи для обновления (с учетом USD): {articles_for_update}")

    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    logger.info(f"Подключение к БД для установки pipeline: {DB_HOST}:{DB_PORT}/{DB_NAME}")

    try:
        cur = conn.cursor()

        # Обновляем pipeline для записей по дате и статьям
        # Преобразуем дату в последний день месяца для сравнения
        last_day_of_month = pd.to_datetime(date_value) + pd.offsets.MonthEnd(0)
        last_day_date = last_day_of_month.date()

        if articles_for_update:
            # Если указаны конкретные статьи, обновляем только их для последнего дня месяца
            placeholders = ','.join(['%s'] * len(articles_for_update))
            query = f"""
            UPDATE {table}
            SET pipeline = %s
            WHERE DATE(\"{db_date_column}\") = %s
            AND article IN ({placeholders})
            """
            cur.execute(query, [pipeline_value, last_day_date] + articles_for_update)
        else:
            # Обновляем все записи для последнего дня месяца
            query = f"""
            UPDATE {table}
            SET pipeline = %s
            WHERE DATE(\"{db_date_column}\") = %s
            """
            cur.execute(query, (pipeline_value, last_day_date))

        rows_affected = cur.rowcount
        logger.info(f"Выполнен UPDATE в таблице {table}: {rows_affected} строк обновлено")

        conn.commit()
        logger.info("Изменения pipeline зафиксированы в БД")
        cur.close()

    except Exception as e:
        logger.error(f"Ошибка при установке pipeline колонки: {e}", exc_info=True)
        conn.rollback()
        logger.info("Транзакция отменена")
        raise e
    finally:
        conn.close()
        logger.info("Соединение с БД для pipeline закрыто")

from io import BytesIO

def process_exchange_rate_file(file_path: str, rate_column: str = 'Курс', date_column: str = 'Дата', sheet_name: str = None):
    """
    Обрабатывает файл с курсами валют и сохраняет их в таблицу EXCHANGE_RATE_TABLE.
    Полностью дропает таблицу и создает заново при каждой загрузке.
    
    Args:
        file_path: путь к Excel файлу с курсами валют
        rate_column: название столбца с курсом (по умолчанию 'Курс')
        date_column: название столбца с датой (по умолчанию 'Дата')
        sheet_name: название листа Excel (если None, берется первый лист)
    
    Returns:
        dict: результат операции с количеством обработанных записей
    """
    logger.info(f"Начало обработки файла курсов валют: {file_path}")
    
    # Читаем Excel файл
    if sheet_name:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        logger.info(f"Загружен лист '{sheet_name}' из Excel файла: {df.shape} (строки x колонки)")
    else:
        df = pd.read_excel(file_path)
        logger.info(f"Загружен Excel файл с курсами: {df.shape} (строки x колонки)")
    
    # Транспонируем DataFrame (аналогично datahandler.py)
    df = (
        df.T
        .reset_index()
        .pipe(lambda x: x.set_axis(x.iloc[0], axis=1))
        .iloc[1:]
        .reset_index(drop=True)
    )
    logger.info(f"DataFrame после транспонирования: {df.shape}")
    logger.info(f"Колонки после транспонирования: {list(df.columns)}")
    
    # Проверяем наличие нужных колонок
    if rate_column not in df.columns:
        available_columns = list(df.columns)
        logger.error(f"Колонка '{rate_column}' не найдена. Доступные колонки: {available_columns}")
        raise ValueError(f"Колонка '{rate_column}' не найдена в файле. Доступные: {available_columns}")
    if date_column not in df.columns:
        available_columns = list(df.columns)
        logger.error(f"Колонка '{date_column}' не найдена. Доступные колонки: {available_columns}")
        raise ValueError(f"Колонка '{date_column}' не найдена в файле. Доступные: {available_columns}")
    
    # Оставляем только нужные колонки
    exchange_data = df[[date_column, rate_column]].copy()
    logger.info(f"Выбраны колонки для курса: {list(exchange_data.columns)}")
    logger.info(f"Первые 3 строки данных курса:")
    logger.info(f"{exchange_data.head(3).to_string()}")
    
    # Преобразуем типы данных
    exchange_data[date_column] = pd.to_datetime(exchange_data[date_column], errors='coerce')
    exchange_data[rate_column] = pd.to_numeric(exchange_data[rate_column], errors='coerce')
    
    # Удаляем строки с некорректными данными
    initial_rows = len(exchange_data)
    exchange_data = exchange_data.dropna()
    final_rows = len(exchange_data)
    
    if initial_rows != final_rows:
        logger.warning(f"Удалено {initial_rows - final_rows} строк с некорректными данными")
    
    logger.info(f"Подготовлено {len(exchange_data)} записей курса для сохранения")
    
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    logger.info(f"Подключение к БД для курсов установлено: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    
    try:
        cur = conn.cursor()
        
        # Проверяем существование таблицы курсов
        cur.execute(f"""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = '{EXCHANGE_RATE_TABLE}'
            )
        """)
        table_exists = cur.fetchone()[0]
        
        if table_exists:
            # Если таблица существует, дропаем её
            cur.execute(f'DROP TABLE {EXCHANGE_RATE_TABLE}')
            logger.info(f"Таблица {EXCHANGE_RATE_TABLE} удалена")
        else:
            logger.info(f"Таблица {EXCHANGE_RATE_TABLE} не существовала")
        
        # Создаем таблицу курсов заново
        create_table_sql = f"""
        CREATE TABLE {EXCHANGE_RATE_TABLE} (
            date DATE NOT NULL PRIMARY KEY,
            exchange_rate DOUBLE PRECISION NOT NULL
        );
        """
        cur.execute(create_table_sql)
        logger.info(f"Таблица {EXCHANGE_RATE_TABLE} создана заново")
        
        # Вставляем курсы в таблицу (без указания id и created_at)
        inserted_rates = 0
        for _, row in exchange_data.iterrows():
            rate_value = row[rate_column]
            date_value = row[date_column].date()  # Конвертируем в date
            
            insert_query = f"""
            INSERT INTO {EXCHANGE_RATE_TABLE} (date, exchange_rate)
            VALUES (%s, %s)
            """
            try:
                cur.execute(insert_query, (date_value, rate_value))
                inserted_rates += 1
            except Exception as e:
                logger.error(f"Ошибка вставки курса для даты {date_value}: {e}")
                continue
        
        logger.info(f"Вставлено {inserted_rates} курсов в таблицу {EXCHANGE_RATE_TABLE}")
        conn.commit()
        logger.info("Курсы валют успешно сохранены")
        
        cur.close()
        
        return {
            "status": "success", 
            "processed_rates": inserted_rates,
            "message": f"Обработано {inserted_rates} курсов валют"
        }
        
    except Exception as e:
        logger.error(f"Ошибка при обработке курсов валют: {e}", exc_info=True)
        conn.rollback()
        logger.info("Транзакция курсов отменена")
        raise e
    finally:
        conn.close()
        logger.info("Соединение с БД для курсов закрыто")

def process_adjustments_file(file_path: str, date_column: str = 'Дата'):
    """
    Обрабатывает файл корректировок и сохраняет их в отдельную таблицу ADJUSTMENTS_TABLE.
    Полностью дропает таблицу и создает заново при каждой загрузке.
    
    Args:
        file_path: путь к Excel файлу с корректировками
        date_column: имя столбца с датой в Excel (например, 'Дата')
    """
    logger.info(f"Начало обработки файла корректировок: {file_path}")
    
    # Читаем файл корректировок
    df_adjustments = pd.read_excel(file_path, sheet_name='Корректировки')
    logger.info(f"Загружен лист 'Корректировки', строк: {len(df_adjustments)}")
    
    # Проверяем обязательные столбцы
    required_columns = ['Статья', 'Год', 'Месяц', 'Корректировка, руб']
    missing_columns = [col for col in required_columns if col not in df_adjustments.columns]
    if missing_columns:
        logger.error(f"Отсутствуют обязательные столбцы: {missing_columns}")
        raise ValueError(f"Отсутствуют обязательные столбцы в файле корректировок: {missing_columns}")
    
    logger.info(f"Все обязательные столбцы присутствуют: {required_columns}")
    
    # Проверяем наличие дополнительных столбцов
    optional_columns = ['Тип', 'Описание']
    for col in optional_columns:
        if col not in df_adjustments.columns:
            df_adjustments[col] = None  # Добавляем пустой столбец если отсутствует
            logger.info(f"Добавлен пустой столбец '{col}'")
    
    # Очищаем данные от пустых строк
    df_adjustments = df_adjustments.dropna(subset=['Статья', 'Год', 'Месяц', 'Корректировка, руб'])
    logger.info(f"После очистки от пустых строк: {len(df_adjustments)} записей")
    
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    logger.info(f"Подключение к БД для корректировок установлено: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    
    try:
        cur = conn.cursor()
        
        # Дропаем таблицу корректировок, если она существует
        cur.execute(f'DROP TABLE IF EXISTS {ADJUSTMENTS_TABLE}')
        logger.info(f"Таблица {ADJUSTMENTS_TABLE} удалена (если существовала)")
        
        # Создаем таблицу корректировок заново
        create_table_sql = f"""
        CREATE TABLE {ADJUSTMENTS_TABLE} (
            "article" TEXT NOT NULL,
            "year" INTEGER NOT NULL,
            "month" INTEGER NOT NULL,
            "adjustment_value" DOUBLE PRECISION NOT NULL,
            "type" TEXT,
            "description" TEXT,
            PRIMARY KEY ("article", "year", "month")
        );
        """
        cur.execute(create_table_sql)
        logger.info(f"Таблица {ADJUSTMENTS_TABLE} создана заново")
        
        # Вставляем корректировки в таблицу
        inserted_adjustments = 0
        for _, row in df_adjustments.iterrows():
            article = row['Статья']
            year = int(row['Год'])
            month = int(row['Месяц'])
            adjustment_value = float(row['Корректировка, руб'])
            adjustment_type = row.get('Тип')
            description = row.get('Описание')
            
            insert_query = f"""
            INSERT INTO {ADJUSTMENTS_TABLE} ("article", "year", "month", "adjustment_value", "type", "description")
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT ("article", "year", "month") 
            DO UPDATE SET 
                "adjustment_value" = EXCLUDED."adjustment_value",
                "type" = EXCLUDED."type",
                "description" = EXCLUDED."description"
            """
            cur.execute(insert_query, (article, year, month, adjustment_value, adjustment_type, description))
            inserted_adjustments += 1
        
        logger.info(f"Вставлено {inserted_adjustments} корректировок в таблицу {ADJUSTMENTS_TABLE}")
        conn.commit()
        logger.info("Корректировки успешно сохранены")
        
        cur.close()
        
        return {
            "status": "success", 
            "processed_adjustments": len(df_adjustments),
            "message": f"Обработано {len(df_adjustments)} корректировок в отдельной таблице"
        }
        
    except Exception as e:
        logger.error(f"Ошибка при обработке корректировок: {e}", exc_info=True)
        conn.rollback()
        logger.info("Транзакция корректировок отменена")
        raise e
    finally:
        conn.close()
        logger.info("Соединение с БД для корректировок закрыто")

def export_pipeline_tables_to_excel(sheet_to_table: dict, make_final_prediction: bool = False, pipeline: str = None) -> BytesIO:
    """
    Экспортирует таблицы pipeline в Excel. Если указан pipeline ('base' или 'base+'), фильтрует строки по этому значению в столбце pipeline.
    """
    dataframes_dict = collect_pipeline_tables_data(sheet_to_table, make_final_prediction, pipeline)
    return dataframes_to_excel(dataframes_dict)


def collect_pipeline_tables_data(sheet_to_table: dict, make_final_prediction: bool = False, pipeline: str = None) -> dict:
    """
    Собирает все таблицы из sheet_to_table и связанные данные в dict датафреймов.
    Если указан pipeline, фильтрует строки по этому значению в столбце pipeline (если такой столбец есть в таблице).
    """
    logger = setup_custom_logging()
    logger.info(f"Начало сбора таблиц. Таблицы: {list(sheet_to_table.keys())}, make_final_prediction: {make_final_prediction}")
    import psycopg2
    import pandas as pd
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    logger.info(f"Подключение к БД для сбора установлено: {DB_HOST}:{DB_PORT}/{DB_NAME}")
    config = load_config(CONFIG_PATH)
    model_article = config.get('model_article', {})
    dataframes_dict = {}
    try:
        # Финальный прогноз
        if make_final_prediction:
            logger.info("Создание финального предсказания")
            if pipeline:
                df_data = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE} WHERE pipeline = %s', conn, params=(pipeline,))
            else:
                df_data = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE}', conn)
            logger.info(f"Загружены данные из {DATA_TABLE}: {len(df_data)} строк")
            try:
                df_adjustments = pd.read_sql_query(f'SELECT "article", "year", "month", "adjustment_value", "description" FROM {ADJUSTMENTS_TABLE}', conn)
                logger.info(f"Загружены корректировки из {ADJUSTMENTS_TABLE}: {len(df_adjustments)} строк")
                df_adjustments['date'] = pd.to_datetime(
                    df_adjustments['year'].astype(str) + '-' + 
                    df_adjustments['month'].astype(str).str.zfill(2) + '-01'
                ) + pd.offsets.MonthEnd(0)
                if 'date' in df_data.columns:
                    df_data['date'] = pd.to_datetime(df_data['date'])
                    df_adjustments['date'] = pd.to_datetime(df_adjustments['date'])
                    logger.info(f"Типы дат приведены к единому формату: df_data.date = {df_data['date'].dtype}, df_adjustments.date = {df_adjustments['date'].dtype}")
                df_data = df_data.merge(
                    df_adjustments[['date', 'article', 'adjustment_value', 'description']], 
                    left_on=['date', 'article'], 
                    right_on=['date', 'article'], 
                    how='left'
                )
                df_data['adjustment_value'] = df_data['adjustment_value'].fillna(0)
                df_data['description'] = df_data['description'].fillna('')
                logger.info(f"Данные объединены с корректировками: {len(df_data)} строк, корректировок с ненулевыми значениями: {(df_data['adjustment_value'] != 0).sum()}")
            except Exception as adj_error:
                logger.warning(f"Не удалось загрузить корректировки: {adj_error}")
                df_data['adjustment_value'] = 0
                df_data['description'] = ''
                logger.info("Добавлены нулевые корректировки")
            original_columns = list(df_data.columns)
            df_data.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df_data.columns]
            mapped_columns = list(df_data.columns)
            logger.info(f"Применено обратное маппинг колонок: {len(original_columns)} -> {len(mapped_columns)}")
            final_rows = []
            # Для каждой уникальной (дата, статья) ищем строку с pipeline из config и берём нужную predict-колонку
            key_cols = ['Дата', 'Статья']
            if 'Дата' not in df_data.columns or 'Статья' not in df_data.columns:
                logger.error('Нет нужных колонок Дата/Статья в данных для финального прогноза')
                df_final = pd.DataFrame(final_rows, columns=['Дата', 'Статья', 'Fact', 'Прогноз', 'Корректировка', 'Финальный прогноз', 'Описание', 'Модель'])
                dataframes_dict['final_prediction'] = df_final
            else:
                # Группируем по (Дата, Статья)
                for (date, article), group in df_data.groupby(['Дата', 'Статья']):
                    model_cfg = model_article.get(article)
                    model_name = None
                    pipeline_cfg = None
                    if isinstance(model_cfg, dict):
                        model_name = model_cfg.get('model')
                        pipeline_cfg = model_cfg.get('pipeline')
                    elif isinstance(model_cfg, str):
                        model_name = model_cfg
                        pipeline_cfg = None
                    if not model_name:
                        continue
                    # Ищем строку с нужным pipeline
                    if pipeline_cfg:
                        row = group[group['pipeline'] == pipeline_cfg].head(1)
                    else:
                        row = group.head(1)
                    if row.empty:
                        continue
                    row = row.iloc[0]
                    pred_col = f'predict_{model_name}'
                    mapped_pred_col = None
                    for col in df_data.columns:
                        if col.lower() == pred_col.lower():
                            mapped_pred_col = col
                            break
                    if not mapped_pred_col:
                        continue
                    adjustment = row.get('Корректировка, руб', 0) or 0
                    description = row.get('Описание', '') or ''
                    final_rows.append({
                        'Дата': date,
                        'Статья': article,
                        'Fact': row.get('Fact'),
                        'Прогноз': row.get(mapped_pred_col),
                        'Корректировка': adjustment,
                        'Финальный прогноз': (row.get(mapped_pred_col) or 0) + (adjustment or 0),
                        'Описание': description,
                        'Модель': model_name
                    })
                df_final = pd.DataFrame(final_rows, columns=['Дата', 'Статья', 'Fact', 'Прогноз', 'Корректировка', 'Финальный прогноз', 'Описание', 'Модель'])
                dataframes_dict['final_prediction'] = df_final
        # Остальные листы
        for sheet, table in sheet_to_table.items():
            if table is None:
                logger.warning(f"Пропуск листа {sheet}: таблица не указана")
                continue
            logger.info(f"Экспорт таблицы {table} в лист {sheet}")
            # Проверяем, есть ли колонка pipeline в таблице
            cur = conn.cursor()
            cur.execute(f"SELECT column_name FROM information_schema.columns WHERE table_name = %s AND column_name = 'pipeline'", (table,))
            has_pipeline = cur.fetchone() is not None
            if pipeline and has_pipeline:
                df = pd.read_sql_query(f'SELECT * FROM {table} WHERE pipeline = %s', conn, params=(pipeline,))
            else:
                df = pd.read_sql_query(f'SELECT * FROM {table}', conn)
            original_columns = list(df.columns)
            df.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df.columns]
            mapped_columns = list(df.columns)
            dataframes_dict[sheet] = df
            logger.info(f"Лист {sheet} экспортирован: {len(df)} строк, колонки: {len(original_columns)} -> {len(mapped_columns)}")
        # Корректировки
        try:
            df_adjustments = pd.read_sql_query(f'SELECT "article", "year", "month", "adjustment_value" FROM {ADJUSTMENTS_TABLE}', conn)
            if not df_adjustments.empty:
                df_adjustments['date'] = pd.to_datetime(
                    df_adjustments['year'].astype(str) + '-' + 
                    df_adjustments['month'].astype(str).str.zfill(2) + '-01'
                ) + pd.offsets.MonthEnd(0)
                df_adjustments = df_adjustments[['date', 'article', 'adjustment_value']]
                logger.info(f"Экспорт корректировок: {len(df_adjustments)} строк")
                df_adjustments.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df_adjustments.columns]
                dataframes_dict['Корректировки'] = df_adjustments
                logger.info("Лист Корректировки добавлен")
            else:
                logger.info("Таблица корректировок пуста, лист не добавлен")
        except Exception as adj_error:
            logger.warning(f"Не удалось экспортировать корректировки: {adj_error}")
            pass
        # Курсы валют
        try:
            df_exchange_rates = pd.read_sql_query(f'SELECT * FROM {EXCHANGE_RATE_TABLE}', conn)
            if not df_exchange_rates.empty:
                logger.info(f"Экспорт курсов валют: {len(df_exchange_rates)} строк")
                df_exchange_rates.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df_exchange_rates.columns]
                dataframes_dict['Курсы валют'] = df_exchange_rates
                logger.info("Лист Курсы валют добавлен")
            else:
                logger.info("Таблица курсов валют пуста, лист не добавлен")
        except Exception as rate_error:
            logger.warning(f"Не удалось экспортировать курсы валют: {rate_error}")
            pass
        return dataframes_dict
    except Exception as e:
        logger.error(f"Ошибка при сборе таблиц: {e}", exc_info=True)
        raise e
    finally:
        conn.close()
        logger.info("Соединение с БД для сбора закрыто")


def dataframes_to_excel(dataframes_dict: dict) -> BytesIO:
    """
    Сохраняет dict датафреймов в Excel-файл в памяти и возвращает BytesIO.
    """
    from utils.excel_formatter import save_dataframes_to_excel
    import tempfile
    import os
    logger = setup_custom_logging()
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()
    try:
        save_dataframes_to_excel(dataframes_dict, tmp.name)
        with open(tmp.name, 'rb') as f:
            output = BytesIO(f.read())
        output.seek(0)
        logger.info("Excel файл создан в памяти и готов к возврату (dataframes_to_excel)")
        return output
    finally:
        os.remove(tmp.name)

def get_article_stats_excel(article_name: str, pipeline: str = None) -> BytesIO:
    """
    Вычисляет статистику по одной статье и возвращает Excel-файл в памяти.
    """
    try:
        # Загружаем конфиг и проверяем, нужно ли добавлять _USD
        config = load_config(CONFIG_PATH)
        usd_articles = config.get('Статьи для предикта в USD', [])
        article_name_db = article_name
        if article_name in usd_articles:
            article_name_db = f"{article_name}_USD"
        # Создаем SQLAlchemy engine для pandas
        db_url = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
        engine = create_engine(db_url)
        conn = psycopg2.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASSWORD,
            dbname=DB_NAME
        )
        # Фильтрация по pipeline если задан
        if pipeline:
            df = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE} WHERE article = %s AND pipeline = %s', engine, params=(article_name_db, pipeline))
        else:
            df = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE} WHERE article = %s', engine, params=(article_name_db,))
        if df.empty:
            raise ValueError(f"Нет данных для статьи: {article_name} (pipeline={pipeline})")
    except Exception as e:
        import traceback
        logger.error(f"[get_article_stats_excel] EXCEPTION: {e}", exc_info=True)
        raise
    df.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df.columns]
    exclude_cols = {'Дата', 'Статья', 'pipeline'}
    value_cols = [col for col in df.columns if col not in exclude_cols]
    df['Год'] = pd.to_datetime(df['Дата']).dt.year
    periods = {
        '2023': (2023, 2023),
        '2024': (2024, 2024),
        '2025': (2025, 2025),
        '2024-2025': (2024, 2025),
        'all': (df['Год'].min(), df['Год'].max()),
    }
    stats = [
        ('Среднее отклонение по модулю за 2023 г. (ABS, %)', 'abs_mean', '2023'),
        ('Среднее отклонение по модулю за 2024 г. (ABS, %)', 'abs_mean', '2024'),
        ('Среднее отклонение по модулю за 2025 г. (ABS, %)', 'abs_mean', '2025'),
        ('Среднее отклонение по модулю за период 2024-2025 (ABS, %)', 'abs_mean', '2024-2025'),
        ('Среднее отклонение по модулю за весь период расчета (ABS, %)', 'abs_mean', 'all'),
        ('Стандартное отклонение за 2023 г.', 'std', '2023'),
        ('Стандартное отклонение за 2024 г.', 'std', '2024'),
        ('Стандартное отклонение за 2025 г.', 'std', '2025'),
        ('Стандартное отклонение за период 2024-2025', 'std', '2024-2025'),
        ('Минимум 2024-2025', 'min', '2024-2025'),
        ('Максимум 2024-2025', 'max', '2024-2025'),
        ('Больше 5%, шт. за 2023', 'gt5', '2023'),
        ('Больше 5%, шт. за 2024', 'gt5', '2024'),
        ('Больше шт. за 2024-2025', 'gt5', '2024-2025'),
    ]
    pct_cols = [col for col in value_cols if 'отклонение %' in col]
    monetary_cols = [col for col in value_cols if col not in pct_cols]
    rates_df = pd.read_sql_query(f'SELECT * FROM {EXCHANGE_RATE_TABLE}', engine)
    if set(rates_df.columns) == {"date", "exchange_rate"}:
        rates_df = rates_df.rename(columns={"date": "Дата", "exchange_rate": "Курс"})
    rates_df['Дата'] = pd.to_datetime(rates_df['Дата'])
    rates_df = rates_df.dropna(subset=['Дата', 'Курс'])
    rates_df = rates_df[rates_df['Курс'] > 0]
    rates_map = dict(zip(rates_df['Дата'].dt.date, rates_df['Курс']))
    avg_rate = rates_df['Курс'].mean() if not rates_df.empty else 1.0
    def resolve_rate(date):
        d = pd.to_datetime(date).date()
        return rates_map.get(d, avg_rate)
    is_usd_article = article_name in usd_articles or article_name_db.endswith('_USD')
    base_cur = 'USD' if is_usd_article else 'RUB'
    target_cur = 'RUB' if is_usd_article else 'USD'
    def convert(val, from_cur, to_cur, date):
        if pd.isna(val):
            return val
        if from_cur == to_cur:
            return val
        rate = resolve_rate(date)
        if from_cur == 'RUB' and to_cur == 'USD':
            return val / rate
        elif from_cur == 'USD' and to_cur == 'RUB':
            return val * rate
        return val
    result = {}
    for stat_name, stat_type, period_key in stats:
        y1, y2 = periods[period_key]
        period_df = df[(df['Год'] >= y1) & (df['Год'] <= y2)]
        row = {}
        # For percent columns: compute as before
        for col in pct_cols:
            if 'Дата' not in period_df.columns:
                continue
            try:
                vals = period_df[[col, 'Дата']]
            except Exception:
                continue
            vals = vals.dropna()
            if vals.empty:
                row[col] = np.nan
                continue
            if stat_type == 'abs_mean':
                row[col] = np.mean(np.abs(vals[col]))
            elif stat_type == 'std':
                row[col] = np.std(vals[col])
            elif stat_type == 'min':
                row[col] = np.min(vals[col])
            elif stat_type == 'max':
                row[col] = np.max(vals[col])
            elif stat_type == 'gt5':
                row[col] = (np.abs(vals[col]) > 5).sum()
        # For monetary columns: compute for both base and target currency
        for col in monetary_cols:
            if 'Дата' not in period_df.columns:
                continue
            try:
                vals = period_df[[col, 'Дата']]
            except Exception:
                continue
            vals = vals.dropna()
            if vals.empty:
                row[f"{col} ({base_cur})"] = np.nan
                row[f"{col} ({target_cur})"] = np.nan
                continue
            # Compute for base currency
            if stat_type == 'abs_mean':
                row[f"{col} ({base_cur})"] = np.mean(np.abs(vals[col]))
                # For target currency, convert each value
                converted_vals = [convert(v, base_cur, target_cur, d) for v, d in zip(vals[col], vals['Дата'])]
                row[f"{col} ({target_cur})"] = np.mean(np.abs(converted_vals))
            elif stat_type == 'std':
                row[f"{col} ({base_cur})"] = np.std(vals[col])
                converted_vals = [convert(v, base_cur, target_cur, d) for v, d in zip(vals[col], vals['Дата'])]
                row[f"{col} ({target_cur})"] = np.std(converted_vals)
            elif stat_type == 'min':
                row[f"{col} ({base_cur})"] = np.min(vals[col])
                converted_vals = [convert(v, base_cur, target_cur, d) for v, d in zip(vals[col], vals['Дата'])]
                row[f"{col} ({target_cur})"] = np.min(converted_vals)
            elif stat_type == 'max':
                row[f"{col} ({base_cur})"] = np.max(vals[col])
                converted_vals = [convert(v, base_cur, target_cur, d) for v, d in zip(vals[col], vals['Дата'])]
                row[f"{col} ({target_cur})"] = np.max(converted_vals)
            elif stat_type == 'gt5':
                row[f"{col} ({base_cur})"] = np.nan
                row[f"{col} ({target_cur})"] = np.nan
        result[stat_name] = row
    stats_df = pd.DataFrame(result).T
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        stats_df.to_excel(writer, sheet_name='Статистика', index=True)
    output.seek(0)
    return output


def export_article_with_agg_excel(article_name: str, pipeline: str = None) -> BytesIO:
    """
    Возвращает Excel-файл с двумя таблицами:
    1. Все строки по статье (дата, факт, все модели в RUB и USD)
    2. Агрегированная таблица (метрика в первом столбце, далее значения по моделям)
    Между ними — пустая строка.
    """
    
    config = load_config(CONFIG_PATH)
    usd_articles = config.get('Статьи для предикта в USD', [])
    article_name_db = article_name
    if article_name in usd_articles:
        article_name_db = f"{article_name}_USD"
    db_url = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    engine = create_engine(db_url)
    # Основная таблица
    if pipeline:
        df = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE} WHERE article = %s AND pipeline = %s', engine, params=(article_name_db, pipeline))
    else:
        df = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE} WHERE article = %s', engine, params=(article_name_db,))
    if df.empty:
        raise ValueError(f"Нет данных для статьи: {article_name} (pipeline={pipeline})")
    df.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df.columns]
    # Курс
    rates_df = pd.read_sql_query(f'SELECT * FROM {EXCHANGE_RATE_TABLE}', engine)
    if set(rates_df.columns) == {"date", "exchange_rate"}:
        rates_df = rates_df.rename(columns={"date": "Дата", "exchange_rate": "Курс"})
    rates_df['Дата'] = pd.to_datetime(rates_df['Дата'])
    rates_df = rates_df.dropna(subset=['Дата', 'Курс'])
    rates_df = rates_df[rates_df['Курс'] > 0]
    rates_map = dict(zip(rates_df['Дата'].dt.date, rates_df['Курс']))
    avg_rate = rates_df['Курс'].mean() if not rates_df.empty else 1.0
    def resolve_rate(date):
        d = pd.to_datetime(date).date()
        return rates_map.get(d, avg_rate)
    # Формируем таблицу: дата, факт (RUB, USD), все predict_* (RUB, USD)
    out_rows = []
    for _, row in df.iterrows():
        date = row.get('Дата')
        fact = row.get('Fact')
        rate = resolve_rate(date)
        fact_rub = fact if not (article_name in usd_articles or article_name_db.endswith('_USD')) else fact * rate
        fact_usd = fact if (article_name in usd_articles or article_name_db.endswith('_USD')) else fact / rate
        out_row = {
            'Дата': date,
            'Факт (RUB)': fact_rub,
            'Факт (USD)': fact_usd,
        }
        # predict_* обработка: если есть %, сохраняем только оригинал; если нет %, сохраняем в RUB и USD
        for col in df.columns:
            if col.startswith('predict_'):
                if '%' in col:
                    out_row[col] = row.get(col)
                else:
                    val = row.get(col)
                    out_row[f'{col} (RUB)'] = val if not (article_name in usd_articles or article_name_db.endswith('_USD')) else val * rate
                    out_row[f'{col} (USD)'] = val if (article_name in usd_articles or article_name_db.endswith('_USD')) else val / rate
        # остальные отклонения % (например, если есть не-predict_ колонки с %)
        for col in df.columns:
            if (
                '%' in col
                and col not in out_row
                and '(RUB)' not in col
                and '(USD)' not in col
            ):
                out_row[col] = row.get(col)
        # добавляем курс доллара в конец
        out_row['Курс доллара'] = rate
        out_rows.append(out_row)
    df_main = pd.DataFrame(out_rows)
    # Получаем агрегированную таблицу через get_article_stats_excel
    agg_excel = get_article_stats_excel(article_name, pipeline=pipeline)
    agg_excel.seek(0)
    with pd.ExcelFile(agg_excel) as xls:
        stats_df = pd.read_excel(xls, sheet_name='Статистика', index_col=0)
    # Итоговый Excel: сначала все данные, потом агрегированная инфа
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # Верхняя таблица
        df_main.to_excel(writer, sheet_name='Статья+Агрегация', index=False, startrow=0)
        # Нижняя таблица (агрегация)
        startrow_agg = len(df_main) + 3
        stats_df_for_write = stats_df.copy()
        stats_df_for_write.insert(0, 'Дата', stats_df_for_write.index)
        stats_df_for_write.reset_index(drop=True, inplace=True)
        desired_columns = list(df_main.columns)
        for col in desired_columns:
            if col not in stats_df_for_write.columns:
                stats_df_for_write[col] = np.nan
        stats_df_for_write = stats_df_for_write[['Дата'] + desired_columns[1:]]
        # Пишем агрегацию (на строку ниже)
        stats_df_for_write.to_excel(writer, sheet_name='Статья+Агрегация', index=False, startrow=startrow_agg+1)
        # Далее — все кастомные стили для нижней таблицы (как было)
        workbook = writer.book
        worksheet = writer.sheets['Статья+Агрегация']
        config = load_config(CONFIG_PATH)
        # --- pipeline-aware model extraction ---
        def get_model_for_article(cfg_dict, article, pipeline):
            val = cfg_dict.get(article)
            if isinstance(val, dict):
                if val.get('pipeline') == pipeline:
                    return val.get('model')
                else:
                    return None
            elif isinstance(val, str):
                return val
            return None

        main_model = get_model_for_article(config.get('model_article', {}), article_name, pipeline)
        reserved_model = get_model_for_article(config.get('model_article_reserved_model', {}), article_name, pipeline)
        all_cols = list(stats_df_for_write.columns)
        pct_col_idxs = [i for i, col in enumerate(all_cols) if '%' in col]
        label_row = [''] * len(all_cols)
        yellow_fmt = workbook.add_format({'bg_color': '#FFEB84'})
        for i in pct_col_idxs:
            col = all_cols[i]
            if main_model and main_model.lower() in col.lower():
                label_row[i] = 'целевая'
                worksheet.write(startrow_agg, i, 'целевая', yellow_fmt)
            elif reserved_model and reserved_model.lower() in col.lower():
                label_row[i] = 'резервная'
                worksheet.write(startrow_agg, i, 'резервная', yellow_fmt)
        # Условное форматирование для процентных столбцов в строках с ABS и всеми "Больше 5%" и "Больше шт." (3-цветная шкала)
        highlight_rows = []
        for i, val in enumerate(stats_df_for_write['Дата']):
            if isinstance(val, str) and (
                'ABS' in val
                or 'Больше 5%' in val
                or 'Больше шт.' in val
            ):
                highlight_rows.append(i)
        pct_cols = [j for j, col in enumerate(stats_df_for_write.columns) if '%' in col]
        def colnum_string(n):
            string = ""
            while n >= 0:
                string = chr(n % 26 + ord('A')) + string
                n = n // 26 - 1
            return string
        for row in highlight_rows:
            row_idx = startrow_agg + 2 + row  # +2 из-за вставленной строки
            if not pct_cols:
                continue
            first_col = pct_cols[0]
            last_col = pct_cols[-1]
            col_letter_start = colnum_string(first_col)
            col_letter_end = colnum_string(last_col)
            cell_range = f'{col_letter_start}{row_idx + 1}:{col_letter_end}{row_idx + 1}'
            worksheet.conditional_format(cell_range, {
                'type': '3_color_scale',
                'min_type': 'min',
                'mid_type': 'percentile',
                'mid_value': 50,
                'max_type': 'max',
                'min_color': '#006100',   # темно-зеленый
                'mid_color': '#FFEB84',   # желтый
                'max_color': '#9C0006',   # красный
            })
    output.seek(0)

    # Форматируем только верхнюю таблицу (df_main) через openpyxl
    from openpyxl import load_workbook
    output2 = BytesIO(output.read())
    wb = load_workbook(output2)
    ws = wb['Статья+Агрегация']
    # Верхняя таблица: строки 1 (заголовок) до len(df_main)+1 включительно
    from openpyxl.styles import numbers, Alignment
    from openpyxl.utils import get_column_letter
    header_height = 40
    ws.row_dimensions[1].height = header_height
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col_idx in range(1, len(df_main.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.alignment = header_alignment
    # Определение типов столбцов для верхней таблицы
    date_cols = []
    pct_cols = []
    num_cols = []
    for col_idx, col_name in enumerate(df_main.columns, start=1):
        col_dtype = df_main[col_name].dtype
        if pd.api.types.is_datetime64_any_dtype(df_main[col_name]):
            date_cols.append(col_idx)
        elif '%' in col_name:
            pct_cols.append(col_idx)
        elif pd.api.types.is_numeric_dtype(col_dtype):
            num_cols.append(col_idx)
    # Форматирование ячеек верхней таблицы
    for row in ws.iter_rows(min_row=2, max_row=len(df_main)+1, min_col=1, max_col=len(df_main.columns)):
        for cell in row:
            col_idx = cell.column
            if col_idx in date_cols and cell.value:
                cell.number_format = numbers.FORMAT_DATE_DDMMYY
            elif col_idx in pct_cols and isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_PERCENTAGE_00
            elif col_idx in num_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '### ### ### ### ##0'
    # Ширина столбцов
    for col_letter in range(1, len(df_main.columns) + 1):
        ws.column_dimensions[get_column_letter(col_letter)].width = 15
    # Автофильтр только для верхней таблицы
    if len(df_main) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df_main.columns))}{len(df_main)+1}"

    # Очищаем строку-заголовок нижней таблицы до применения форматирования
    header_row_idx = startrow_agg + 2
    for cell in ws[header_row_idx]:
        cell.value = None
        cell.border = None

    # После очистки строка с данными становится на место бывшей строки заголовка
    # Форматируем проценты в нижней таблице (теперь первая строка данных — header_row_idx+1)
    agg_data_start = header_row_idx + 1
    agg_data_end = ws.max_row
    agg_pct_cols = [i+1 for i, col in enumerate(stats_df_for_write.columns) if '%' in col]
    # Получаем список строк, для которых НЕ нужно процентное форматирование ("Больше 5%" и "Больше шт.")
    skip_percent_rows = []
    for i, val in enumerate(stats_df_for_write['Дата']):
        if isinstance(val, str) and (val.startswith('Больше 5%') or val.startswith('Больше шт.')):
            skip_percent_rows.append(i)
    # Определим индексы всех числовых столбцов (не процентов)
    agg_num_cols = [i+1 for i, col in enumerate(stats_df_for_write.columns) if '%' not in col]
    for row_idx, row in enumerate(ws.iter_rows(min_row=agg_data_start, max_row=agg_data_end, min_col=1, max_col=len(stats_df_for_write.columns))):
        for cell in row:
            # Если это строка "Больше 5%" или "Больше шт.", не применять процентный формат и привести к int
            if row_idx in skip_percent_rows:
                if cell.column in agg_pct_cols and isinstance(cell.value, (int, float)):
                    cell.value = int(round(cell.value))
            else:
                if cell.column in agg_pct_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = "0%"
            # Для всех непроцентных числовых столбцов — округление до целого
            if cell.column in agg_num_cols and isinstance(cell.value, (int, float)):
                cell.value = int(round(cell.value))

    output_final = BytesIO()
    wb.save(output_final)
    output_final.seek(0)
    return output_final

def get_bullet_chart_data() -> dict:
    import sqlalchemy
    db_url = f"postgresql+psycopg2://{DB_USER}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
    engine = sqlalchemy.create_engine(db_url)
    conn = psycopg2.connect(
        host=DB_HOST,
        port=DB_PORT,
        user=DB_USER,
        password=DB_PASSWORD,
        dbname=DB_NAME
    )
    config = load_config(CONFIG_PATH)
    model_article = config.get('model_article', {})
    usd_articles = config.get('Статьи для предикта в USD', [])
    def get_model_and_pipeline_for_article(cfg_dict, article):
        val = cfg_dict.get(article)
        if isinstance(val, dict):
            return val.get('model'), val.get('pipeline')
        elif isinstance(val, str):
            return val, None
        return None, None
    try:
        df = pd.read_sql_query(f'SELECT * FROM {DATA_TABLE}', engine)
        df.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df.columns]
        # Приводим все имена колонок к нижнему регистру для унификации поиска
        df.columns = [col.lower() for col in df.columns]
        # Для каждой (дата, статья) ищем только строку с pipeline==целевой и берем predict_{model}*
        result = []
        # Словарь: (article_for_front, date) -> row
        used_keys = set()
        for idx, row in df.iterrows():
            row_dict = {k.lower(): v for k, v in row.items()}
            article = row_dict.get('статья')
            date = row_dict.get('дата')
            pipeline_val = row_dict.get('pipeline')
            if not article or date is None:
                continue
            # base_article без _USD
            base_article = article[:-4] if article.endswith('_usd') else article
            model, pipeline_cfg = get_model_and_pipeline_for_article(model_article, base_article)
            if not model or not pipeline_cfg:
                continue
            # pipeline должен совпадать с config
            if str(pipeline_val).lower() != str(pipeline_cfg).lower():
                continue
            # Для фронта статья без _USD
            article_for_front = article[:-4] if article.endswith('_usd') else article
            # Ключ для уникальности
            key = (article_for_front, date)
            if key in used_keys:
                continue
            used_keys.add(key)
            # predict_{model} разница и predict_{model} отклонение %
            diff_key = f'predict_{model} разница'.lower()
            dev_key = f'predict_{model} отклонение %'.lower()
            difference = row_dict.get(diff_key)
            deviation = row_dict.get(dev_key)
            if isinstance(date, pd.Timestamp):
                date_str = date.strftime('%Y-%m-%d')
            else:
                date_str = str(date)
            def safe_float(val):
                try:
                    f = float(val)
                    if np.isnan(f) or np.isinf(f):
                        return None
                    return f
                except Exception:
                    return None
            result.append({
                'article': article_for_front,
                'model': model,
                'date': date_str,
                'deviation': safe_float(deviation),
                'difference': safe_float(difference),
                'pipeline': pipeline_val
            })
        # Получить курсы валют
        try:
            if not EXCHANGE_RATE_TABLE:
                raise ValueError('EXCHANGE_RATE_TABLE env var is not set')
            df_rates = pd.read_sql_query(f'SELECT * FROM {EXCHANGE_RATE_TABLE}', engine)
            df_rates.columns = [REVERSE_COLUMN_MAPPING.get(col, col) for col in df_rates.columns]
            exchange_rates = df_rates.to_dict(orient='records')
            # Преобразовать все значения типа date в строку
            for row in exchange_rates:
                for k, v in row.items():
                    if hasattr(v, 'isoformat'):
                        row[k] = v.isoformat()
        except Exception as e:
            logger.warning(f"Не удалось получить курсы валют: {e}")
            exchange_rates = []
        return {'data': result, 'exchange_rates': exchange_rates}
    finally:
        conn.close()